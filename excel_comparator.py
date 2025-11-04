#!/usr/bin/env python3
"""
Excel Comparator Tool (Source vs Org Data)

Author: Yugendran S
License: MIT

Compares two Excel files — Source and Org Data — using a mapping file.
Generates detailed comparison, summary, and mismatch reports in one Excel output.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import math

# ---------------------------
# CONFIG (edit these paths)
# ---------------------------
SOURCE_FILE = "sample_files/Source_File_Account_details.xlsx"
ORG_DATA_FILE = "sample_files/OrgData_Account.xlsx"
MAPPING_FILE = "sample_files/Account_Mapping.xlsx"
OUTPUT_FILE = "comparison_output.xlsx"
ENTITY_NAME = "Object - Validation"
SOURCE_ID_COL = "Source ID"
MAX_ROWS_PER_SHEET = 1_000_000

# ---------------------------
# Helper functions
# ---------------------------
FALSE_SET = {"", "0", "NO", "N", "FALSE", "NONE", "NULL"}
TRUE_SET = {"1", "YES", "Y", "TRUE"}

def read_all_sheets(file_path):
    """Reads all sheets from Excel file and combines them into one DataFrame."""
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    df_list = []
    for sheet in xls.sheet_names:
        temp_df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl", na_filter=False)
        temp_df["__sheet_name__"] = sheet
        df_list.append(temp_df)
    return pd.concat(df_list, ignore_index=True) if df_list else pd.DataFrame()

def normalize_bool(val):
    """Converts boolean-like strings to 'TRUE'/'FALSE'."""
    if pd.isna(val):
        v = ""
    else:
        v = str(val).strip().upper()
    if v in FALSE_SET:
        return "FALSE"
    if v in TRUE_SET:
        return "TRUE"
    return v

def sanitize_sheet_name(name):
    invalid_chars = ['\\','/','*','?',':','[',']']
    for c in invalid_chars:
        name = name.replace(c, '_')
    return str(name)[:31]

def contains_flag(src_val, org_val):
    """Checks if Org Data value contains Source value (case-insensitive)."""
    src_norm = normalize_bool(src_val)
    org_norm = normalize_bool(org_val)
    if src_norm == org_norm:
        return True
    s_src = "" if pd.isna(src_val) else str(src_val).strip()
    s_org = "" if pd.isna(org_val) else str(org_val).strip()
    if s_src == "" and s_org == "":
        return True
    try:
        return (s_src != "") and (s_src.lower() in s_org.lower())
    except Exception:
        return False

# ---------------------------
# MAIN PROCESS
# ---------------------------

def main():
    print("Reading files...")
    source_df = read_all_sheets(SOURCE_FILE)
    orgdata_df = read_all_sheets(ORG_DATA_FILE)
    mapping_df = pd.read_excel(MAPPING_FILE, engine="openpyxl", dtype=str)

    if SOURCE_ID_COL not in source_df.columns or SOURCE_ID_COL not in orgdata_df.columns:
        raise ValueError(f"'{SOURCE_ID_COL}' not found in one of the input files.")

    source_df[SOURCE_ID_COL] = source_df[SOURCE_ID_COL].astype(str).str.strip()
    orgdata_df[SOURCE_ID_COL] = orgdata_df[SOURCE_ID_COL].astype(str).str.strip()

    source_df = source_df.drop_duplicates(subset=[SOURCE_ID_COL], keep="first").set_index(SOURCE_ID_COL)
    orgdata_df = orgdata_df.drop_duplicates(subset=[SOURCE_ID_COL], keep="first").set_index(SOURCE_ID_COL)

    mapping_pairs = list(zip(
        mapping_df["Source_Column"].astype(str).str.strip(),
        mapping_df["OrgData_Column"].astype(str).str.strip()
    ))
    valid_mapping_pairs = [(src, org) for src, org in mapping_pairs if src in source_df.columns and org in orgdata_df.columns]

    print(f"✅ {len(valid_mapping_pairs)} valid mapped field pairs detected.")
    intersection_ids = source_df.index.intersection(orgdata_df.index)
    print(f"🔗 Found {len(intersection_ids)} common Source IDs.")

    comparison_data = []
    failed_records = {}
    summary_counts = {}

    for src_col, org_col in valid_mapping_pairs:
        match_count = 0
        mismatch_count = 0
        failed_records[(src_col, org_col)] = []

        for src_id in intersection_ids:
            src_val = source_df.at[src_id, src_col] if src_col in source_df.columns else None
            org_val = orgdata_df.at[src_id, org_col] if org_col in orgdata_df.columns else None
            norm_src = normalize_bool(src_val)
            norm_org = normalize_bool(org_val)
            if (str(src_val).strip() == "" and str(org_val).strip() == ""):
                is_match = True
            elif norm_src == norm_org:
                is_match = True
            else:
                is_match = contains_flag(src_val, org_val)
            if is_match:
                match_count += 1
            else:
                mismatch_count += 1
                failed_records[(src_col, org_col)].append({
                    "Source ID": src_id,
                    f"{src_col} (Source)": src_val,
                    f"{org_col} (OrgData)": org_val,
                })
            comparison_data.append({
                "Source ID": src_id,
                f"{src_col} (Source)": src_val,
                f"{org_col} (OrgData)": org_val,
                f"{src_col} (Match)": is_match
            })
        summary_counts[(src_col, org_col)] = {"total": len(intersection_ids), "match": match_count, "fail": mismatch_count}

    summary_records = []
    for (src_col, org_col), counts in summary_counts.items():
        total = counts["total"]
        matched = counts["match"]
        failed = counts["fail"]
        outcome = "Pass" if matched == total and total > 0 else "Fail"
        summary_records.append({
            "Entity - Cycle": ENTITY_NAME,
            "Source Field": src_col,
            "OrgData Field": org_col,
            "Total Records": total,
            "Matched Count": matched,
            "Mismatched Count": failed,
            "Outcome": outcome
        })

    summary_df = pd.DataFrame(summary_records)
    if not summary_df.empty:
        total_row = {
            "Entity - Cycle": ENTITY_NAME,
            "Source Field": "TOTAL",
            "OrgData Field": "",
            "Total Records": int(summary_df["Total Records"].sum()),
            "Matched Count": int(summary_df["Matched Count"].sum()),
            "Mismatched Count": int(summary_df["Mismatched Count"].sum()),
            "Outcome": ""
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)

    comparison_df = pd.DataFrame(comparison_data)

    print("📤 Writing output...")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        comparison_df.to_excel(writer, sheet_name="Comparison", index=False)

        for (src_col, org_col), group in failed_records.items():
            if not group:
                continue
            group_df = pd.DataFrame(group)
            group_df["OrgData Contains Source"] = group_df.apply(
                lambda r: contains_flag(r[f"{src_col} (Source)"], r[f"{org_col} (OrgData)"]), axis=1
            )
            sheet_name = sanitize_sheet_name(f"{src_col}_vs_{org_col}")
            group_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(OUTPUT_FILE)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        outcome_col = None
        for cell in ws[1]:
            if cell.value == "Outcome":
                outcome_col = cell.col_idx
                break
        if outcome_col:
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=outcome_col, max_col=outcome_col):
                cell = row[0]
                if cell.value == "Pass":
                    cell.fill = green_fill
                elif cell.value == "Fail":
                    cell.fill = red_fill

    wb.save(OUTPUT_FILE)
    print(f"✅ Comparison completed! Output file: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
